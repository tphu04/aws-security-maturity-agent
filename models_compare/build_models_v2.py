"""
Build Models_v2.xlsx — Enhanced model comparison workbook
==========================================================
Tells the story: Survey -> Filter -> Shortlist -> Test -> Benchmark -> Decision

Sheets:
  1. 1_Khao_sat_LLM        — 10 large LLMs (>= 7B) from initial research
  2. 2_Khao_sat_SLM        — Small LLMs (<= 4B) + GGUF community quants
  3. 3_Loc_phan_cung       — Hardware filter (6 GB VRAM budget)
  4. 4_Test_GGUF_ban_dau   — Original GGUF tests (preserved from old Excel)
  5. 5_Benchmark_thuc_te   — Report Agent benchmark (real measurements)
  6. 6_Quyet_dinh          — Weighted decision matrix -> gemma3:4b wins
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT_PATH = Path(__file__).parent / "Models_v2.xlsx"
OLD_PATH = Path(__file__).parent / "Models.xlsx"

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
SUBHEADER_FONT = Font(bold=True, color="000000", size=10)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")  # green
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")  # red
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")  # yellow
CHOSEN_FILL = PatternFill("solid", fgColor="70AD47")  # dark green for winner
CHOSEN_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="808080")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# REFERENCE LINKS (researched)
# ---------------------------------------------------------------------------
REFS = {
    "gpt-oss-20b": {
        "dev": "OpenAI", "date": "2025-08",
        "official": "https://openai.com/index/introducing-gpt-oss/",
        "hf": "https://huggingface.co/openai/gpt-oss-20b",
        "paper": "",
    },
    "Gemma 3 27B IT": {
        "dev": "Google DeepMind", "date": "2025-03",
        "official": "https://blog.google/technology/developers/gemma-3/",
        "hf": "https://huggingface.co/google/gemma-3-27b-it",
        "paper": "https://arxiv.org/abs/2503.19786",
    },
    "Mistral Small 3.1": {
        "dev": "Mistral AI", "date": "2025-03",
        "official": "https://mistral.ai/news/mistral-small-3-1",
        "hf": "https://huggingface.co/mistralai/Mistral-Small-3.1-24B-Instruct-2503",
        "paper": "",
    },
    "Qwen2.5-32B-Instruct": {
        "dev": "Alibaba (Qwen Team)", "date": "2024-09",
        "official": "https://qwenlm.github.io/blog/qwen2.5/",
        "hf": "https://huggingface.co/Qwen/Qwen2.5-32B-Instruct",
        "paper": "https://arxiv.org/abs/2412.15115",
    },
    "DeepSeek-R1": {
        "dev": "DeepSeek-AI", "date": "2025-01",
        "official": "https://api-docs.deepseek.com/news/news250120",
        "hf": "https://huggingface.co/deepseek-ai/DeepSeek-R1",
        "paper": "https://arxiv.org/abs/2501.12948",
    },
    "Gemma 3 12B IT": {
        "dev": "Google DeepMind", "date": "2025-03",
        "official": "https://blog.google/technology/developers/gemma-3/",
        "hf": "https://huggingface.co/google/gemma-3-12b-it",
        "paper": "https://arxiv.org/abs/2503.19786",
    },
    "Llama 3.1 8B Instruct": {
        "dev": "Meta AI", "date": "2024-07",
        "official": "https://ai.meta.com/blog/meta-llama-3-1/",
        "hf": "https://huggingface.co/meta-llama/Llama-3.1-8B-Instruct",
        "paper": "https://arxiv.org/abs/2407.21783",
    },
    "Mistral 7B Instruct": {
        "dev": "Mistral AI", "date": "2023-09",
        "official": "https://mistral.ai/news/announcing-mistral-7b",
        "hf": "https://huggingface.co/mistralai/Mistral-7B-Instruct-v0.3",
        "paper": "https://arxiv.org/abs/2310.06825",
    },
    "Qwen2.5-7B-Instruct": {
        "dev": "Alibaba (Qwen Team)", "date": "2024-09",
        "official": "https://qwenlm.github.io/blog/qwen2.5/",
        "hf": "https://huggingface.co/Qwen/Qwen2.5-7B-Instruct",
        "paper": "https://arxiv.org/abs/2412.15115",
    },
    "Gemma 3 4B IT": {
        "dev": "Google DeepMind", "date": "2025-03",
        "official": "https://blog.google/technology/developers/gemma-3/",
        "hf": "https://huggingface.co/google/gemma-3-4b-it",
        "paper": "https://arxiv.org/abs/2503.19786",
    },
    "Gemma 2 2B IT": {
        "dev": "Google DeepMind", "date": "2024-07",
        "official": "https://blog.google/technology/developers/google-gemma-2/",
        "hf": "https://huggingface.co/google/gemma-2-2b-it",
        "paper": "https://arxiv.org/abs/2408.00118",
    },
    "Llama 3.2 3B Instruct": {
        "dev": "Meta AI", "date": "2024-09",
        "official": "https://ai.meta.com/blog/llama-3-2-connect-2024-vision-edge-mobile-devices/",
        "hf": "https://huggingface.co/meta-llama/Llama-3.2-3B-Instruct",
        "paper": "",
    },
    "Phi-4-mini": {
        "dev": "Microsoft", "date": "2025-02",
        "official": "https://azure.microsoft.com/en-us/blog/empowering-innovation-the-next-generation-of-the-phi-family/",
        "hf": "https://huggingface.co/microsoft/Phi-4-mini-instruct",
        "paper": "https://arxiv.org/abs/2503.01743",
    },
    "Arcee-VyLinh": {
        "dev": "Arcee AI", "date": "2024-11",
        "official": "https://www.arcee.ai/blog/introducing-arcee-vylinh-a-powerful-3b-parameter-vietnamese-language-model",
        "hf": "https://huggingface.co/arcee-ai/Arcee-VyLinh",
        "paper": "",
    },
    "DeepSeek-R1-Distill-Qwen-7B": {
        "dev": "DeepSeek-AI", "date": "2025-01",
        "official": "https://api-docs.deepseek.com/news/news250120",
        "hf": "https://huggingface.co/deepseek-ai/DeepSeek-R1-Distill-Qwen-7B",
        "paper": "https://arxiv.org/abs/2501.12948",
    },
    "bartowski/Llama-3.2-3B-Instruct-GGUF": {
        "dev": "bartowski (community quant)", "date": "2024-09",
        "official": "", "paper": "",
        "hf": "https://huggingface.co/bartowski/Llama-3.2-3B-Instruct-GGUF",
    },
    "bartowski/gemma-2-2b-it-GGUF": {
        "dev": "bartowski (community quant)", "date": "2024-07",
        "official": "", "paper": "",
        "hf": "https://huggingface.co/bartowski/gemma-2-2b-it-GGUF",
    },
    "aws-prototyping/codefu-7b-v0.1": {
        "dev": "AWS Prototyping", "date": "2024",
        "official": "", "paper": "",
        "hf": "https://huggingface.co/aws-prototyping/codefu-7b-v0.1",
    },
    "aws-prototyping/MegaBeam-Mistral-7B-300k": {
        "dev": "AWS Prototyping", "date": "2024",
        "official": "", "paper": "",
        "hf": "https://huggingface.co/aws-prototyping/MegaBeam-Mistral-7B-300k",
    },
    "aws-prototyping/OmniLong-Qwen2.5-VL-7B": {
        "dev": "AWS Prototyping", "date": "2025",
        "official": "", "paper": "",
        "hf": "https://huggingface.co/aws-prototyping/OmniLong-Qwen2.5-VL-7B",
    },
}


# ---------------------------------------------------------------------------
# BENCHMARK DATA (from real measurements in benchmark_results/benchmark_20260413_164412.json)
# ---------------------------------------------------------------------------
BENCHMARK_DATA = {
    "gemma3:4b": {
        "exec_summary":       {"duration": 5.69,  "speed": 69.70, "words": 207, "limit": 400, "vram": 4849, "violations": 0},
        "remediation_detail": {"duration": 10.90, "speed": 70.47, "words": 497, "limit": 350, "vram": 4842, "violations": 1},
        "recommendations":    {"duration": 11.10, "speed": 71.53, "words": 503, "limit": 300, "vram": 4841, "violations": 2},
        "summary": {"avg_speed": 61.8, "total_time": 27.7, "max_vram": 4849, "total_violations": 3},
        "quality": {"vietnamese": 5, "instruction": 3, "accuracy": 5, "tone": 5, "concise": 3},
    },
    "qwen2.5:7b-instruct-q4_K_M": {
        "exec_summary":       {"duration": 20.50, "speed": 11.39, "words": 140, "limit": 400, "vram": 5352, "violations": 0},
        "remediation_detail": {"duration": 25.97, "speed": 11.65, "words": 209, "limit": 350, "vram": 5352, "violations": 0},
        "recommendations":    {"duration": 36.82, "speed": 9.98,  "words": 281, "limit": 300, "vram": 5280, "violations": 0},
        "summary": {"avg_speed": 10.3, "total_time": 83.3, "max_vram": 5352, "total_violations": 0},
        "quality": {"vietnamese": 3, "instruction": 5, "accuracy": 5, "tone": 4, "concise": 5},
    },
    "deepseek-r1:7b": {
        "exec_summary":       {"duration": 78.67, "speed": 11.12, "words": 221, "limit": 400, "vram": 5276, "violations": 0},
        "remediation_detail": {"duration": 86.74, "speed": 12.12, "words": 343, "limit": 350, "vram": 5295, "violations": 0},
        "recommendations":    {"duration": 93.07, "speed": 11.27, "words": 476, "limit": 300, "vram": 5342, "violations": 2},
        "summary": {"avg_speed": 11.2, "total_time": 258.5, "max_vram": 5342, "total_violations": 2},
        "quality": {"vietnamese": 2, "instruction": 3, "accuracy": 4, "tone": 4, "concise": 3},
    },
    "llama3.2:latest": {
        "exec_summary":       {"duration": 6.80, "speed": 104.40, "words": 316, "limit": 400, "vram": 3837, "violations": 5},
        "remediation_detail": {"duration": 5.49, "speed": 98.80,  "words": 202, "limit": 350, "vram": 3842, "violations": 1},
        "recommendations":    {"duration": 5.48, "speed": 94.39,  "words": 318, "limit": 300, "vram": 3856, "violations": 3},
        "summary": {"avg_speed": 83.0, "total_time": 17.8, "max_vram": 3856, "total_violations": 9},
        "quality": {"vietnamese": 2, "instruction": 1, "accuracy": 2, "tone": 1, "concise": 2},
    },
}


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def set_header(ws, row_idx, headers, widths=None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_idx, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def write_row(ws, row_idx, values, fill=None, font=None, bold=False):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col, value=v)
        c.alignment = LEFT
        c.border = BORDER
        if fill:
            c.fill = fill
        if font:
            c.font = font
        elif bold:
            c.font = Font(bold=True)


def add_section_title(ws, row_idx, title, span=10):
    c = ws.cell(row=row_idx, column=1, value=title)
    c.fill = SUBHEADER_FILL
    c.font = SUBHEADER_FONT
    c.alignment = LEFT
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span)


def add_hyperlink(ws, row, col, url, display=None):
    if not url:
        return
    c = ws.cell(row=row, column=col, value=display or url)
    c.hyperlink = url
    c.font = Font(color="0563C1", underline="single")


# ---------------------------------------------------------------------------
# SHEET 1 — KHAO SAT LLM (>= 7B)
# ---------------------------------------------------------------------------
def build_sheet1_khao_sat_llm(wb):
    ws = wb.create_sheet("1_Khao_sat_LLM")
    ws["A1"] = "GIAI ĐOẠN 1 — KHẢO SÁT TOÀN CẢNH LLM"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:L1")

    ws["A2"] = (
        "Mục tiêu: Khảo sát các LLM hiện đại (>= 7B params) để hiểu landscape. "
        "Kết luận: Phần lớn yêu cầu VRAM >= 12 GB — KHÔNG phù hợp với hạ tầng local "
        "(RTX 3060 Laptop 6 GB VRAM). Chuyển sang nhóm SLM (Sheet 2)."
    )
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 45

    headers = [
        "Model", "Developer", "Release", "Params", "Context",
        "MMLU", "GSM8K", "HumanEval", "VRAM Q4 (GB)",
        "License", "Tham khảo chính", "HuggingFace",
    ]
    widths = [28, 22, 10, 10, 12, 10, 12, 12, 14, 28, 45, 50]
    set_header(ws, 4, headers, widths)

    rows = [
        # (model, params, context, mmlu, gsm8k, humaneval, vram_q4, license)
        ("gpt-oss-20b",           "20B",   "128K",  "85.3%",   "~92%",    "~60%",  "~11",  "Apache 2.0"),
        ("Gemma 3 27B IT",        "27B",   "128K",  "76.9%",   "95.9%",   "87.8%", "~16",  "Gemma License"),
        ("Mistral Small 3.1",     "24B",   "128K",  "~81%",    "~91%",    "~68%",  "~14",  "Apache 2.0"),
        ("Qwen2.5-32B-Instruct",  "32B",   "128K",  "83.9%",   "95.9%",   "88.4%", "~19",  "Apache 2.0"),
        ("DeepSeek-R1",           "37B act", "128K",  "90.8%",   "97.3%",   "65.9%", "~22",  "MIT / OpenRAIL"),
        ("Gemma 3 12B IT",        "12B",   "128K",  "71.9%",   "94.4%",   "85.4%", "~7",   "Gemma License"),
        ("Llama 3.1 8B Instruct", "8B",    "128K",  "78.6%",   "92.5%",   "77.3%", "~5",   "Llama 3.1 Community"),
        ("Mistral 7B Instruct",   "7B",    "32K",   "60.1%",   "52.2%",   "30.5%", "~4.5", "Apache 2.0"),
        ("Qwen2.5-7B-Instruct",   "7B",    "128K",  "75.4%",   "91.6%",   "84.8%", "~5",   "Apache 2.0"),
    ]

    for i, (name, params, ctx, mmlu, gsm8k, he, vram, lic) in enumerate(rows, start=5):
        ref = REFS.get(name, {})
        values = [
            name, ref.get("dev", ""), ref.get("date", ""),
            params, ctx, mmlu, gsm8k, he, vram, lic, "", "",
        ]
        write_row(ws, i, values)
        # Add hyperlinks
        add_hyperlink(ws, i, 11, ref.get("official") or ref.get("paper"), ref.get("official") or ref.get("paper"))
        add_hyperlink(ws, i, 12, ref.get("hf"), ref.get("hf"))

    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# SHEET 2 — KHAO SAT SLM (<= 4B)
# ---------------------------------------------------------------------------
def build_sheet2_khao_sat_slm(wb):
    ws = wb.create_sheet("2_Khao_sat_SLM")
    ws["A1"] = "GIAI ĐOẠN 2 — KHẢO SÁT SLM (SMALL LANGUAGE MODEL)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:L1")

    ws["A2"] = (
        "Mục tiêu: Mở rộng candidate bằng các SLM (<= 4B params) và GGUF community quants. "
        "Tiêu chí bổ sung: Hỗ trợ tiếng Việt (bắt buộc cho Report Agent). "
        "Kết luận: Chuyển sang Sheet 3 để lọc theo phần cứng và ngôn ngữ."
    )
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 45

    headers = [
        "Model", "Developer", "Release", "Params", "Context",
        "MMLU", "IFEval", "Vietnamese", "VRAM Q4 (GB)",
        "License", "Tham khảo chính", "HuggingFace",
    ]
    widths = [38, 22, 10, 10, 12, 10, 10, 14, 14, 22, 45, 55]
    set_header(ws, 4, headers, widths)

    # (name, params, context, mmlu, ifeval, vn_support, vram_q4, license)
    rows = [
        # === Official small models ===
        ("Gemma 3 4B IT",           "4B",   "128K", "59.6%",   "90.2%",  "Có (140+ ngôn ngữ)",  "~2.8", "Gemma License"),
        ("Gemma 2 2B IT",           "2B",   "8K",   "51.3%",   "~68%",   "Có",                  "~1.7", "Gemma License"),
        ("Llama 3.2 3B Instruct",   "3B",   "128K", "63.4%",   "~70%",   "Hạn chế",             "~2.0", "Llama 3.2 Community"),
        ("Phi-4-mini",              "3.8B", "128K", "67.3%",   "~75%",   "KHÔNG hỗ trợ",        "~2.5", "MIT"),
        ("Arcee-VyLinh",            "3B",   "8K",   "~55%",    "N/A",    "Chuyên biệt VN",      "~2.0", "Custom"),
        ("DeepSeek-R1-Distill-Qwen-7B", "7B", "128K", "~70%",  "N/A",    "Có nhưng output EN",  "~4.7", "MIT"),
        # === GGUF community quants ===
        ("bartowski/Llama-3.2-3B-Instruct-GGUF",            "3.2B", "128K", "63.4%",  "~70%", "Hạn chế", "~2.0", "Llama 3.2 Community"),
        ("bartowski/gemma-2-2b-it-GGUF",                    "2.6B", "8K",   "51.3%",  "~68%", "Có",      "~1.7", "Gemma License"),
        ("aws-prototyping/codefu-7b-v0.1",                  "7.6B", "128K", "N/A",    "N/A",  "Kém",     "~4.7", "Custom"),
        ("aws-prototyping/MegaBeam-Mistral-7B-300k",        "7.2B", "300K", "~60%",   "N/A",  "Kém",     "~4.4", "Apache 2.0"),
        ("aws-prototyping/OmniLong-Qwen2.5-VL-7B",          "7.6B", "524K", "~75%",   "N/A",  "Có",      "~6.0", "Apache 2.0"),
    ]

    for i, (name, params, ctx, mmlu, ife, vn, vram, lic) in enumerate(rows, start=5):
        ref = REFS.get(name, {})
        values = [
            name, ref.get("dev", ""), ref.get("date", ""),
            params, ctx, mmlu, ife, vn, vram, lic, "", "",
        ]
        write_row(ws, i, values)

        # Highlight VN column
        vn_cell = ws.cell(row=i, column=8)
        if vn.startswith("Có") or vn.startswith("Chuyên"):
            vn_cell.fill = PASS_FILL
        elif vn.startswith("Hạn chế") or vn.startswith("Kém"):
            vn_cell.fill = WARN_FILL
        elif vn.startswith("KHÔNG"):
            vn_cell.fill = FAIL_FILL

        add_hyperlink(ws, i, 11, ref.get("official") or ref.get("paper"), ref.get("official") or ref.get("paper"))
        add_hyperlink(ws, i, 12, ref.get("hf"), ref.get("hf"))

    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# SHEET 3 — HARDWARE FILTER
# ---------------------------------------------------------------------------
def build_sheet3_loc_phan_cung(wb):
    ws = wb.create_sheet("3_Loc_phan_cung")
    ws["A1"] = "GIAI ĐOẠN 3 — LỌC THEO PHẦN CỨNG & TIẾNG VIỆT"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:H1")

    ws["A2"] = (
        "Ngân sách phần cứng: RTX 3060 Laptop — 6 GB VRAM, 16 GB RAM. "
        "Tiêu chí PASS: (1) VRAM Q4 <= 5.5 GB (dư chỗ cho KV cache), "
        "(2) Hỗ trợ tiếng Việt ở mức trung bình trở lên. "
        "Kết quả: 4 candidate cho giai đoạn benchmark thực tế."
    )
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 45

    headers = [
        "Model", "Params", "VRAM Q4 (GB)", "Hardware OK?",
        "Tiếng Việt", "VN OK?", "Kết luận", "Ghi chú",
    ]
    widths = [38, 10, 14, 14, 22, 10, 18, 45]
    set_header(ws, 4, headers, widths)

    rows = [
        # Large LLMs — all fail hardware
        ("Gemma 3 27B IT",          "27B",   16.0,  "Có (140+)",      "Vượt ngân sách VRAM 16 GB"),
        ("Qwen2.5-32B-Instruct",    "32B",   19.0,  "Có (29+)",       "Vượt ngân sách VRAM 19 GB"),
        ("Mistral Small 3.1",       "24B",   14.0,  "Có",             "Vượt ngân sách VRAM 14 GB"),
        ("DeepSeek-R1",             "37B",   22.0,  "Có nhưng EN",    "Vượt ngân sách VRAM 22 GB"),
        ("gpt-oss-20b",             "20B",   11.0,  "Có",             "Vượt ngân sách VRAM 11 GB"),
        ("Gemma 3 12B IT",          "12B",   7.0,   "Có (140+)",      "Vượt ngân sách (sát biên)"),
        # Medium — partial pass
        ("Llama 3.1 8B Instruct",   "8B",    5.0,   "Hạn chế",        "VRAM OK nhưng VN kém"),
        ("Mistral 7B Instruct",     "7B",    4.5,   "Kém",            "VRAM OK nhưng VN kém + MMLU thấp"),
        ("Qwen2.5-7B-Instruct",     "7B",    5.0,   "Có (29+)",       "PASS — vào benchmark"),
        ("DeepSeek-R1-Distill-Qwen-7B", "7B", 4.7,  "Có nhưng EN",    "PASS — vào benchmark"),
        # Small — most pass
        ("Gemma 3 4B IT",           "4B",    2.8,   "Có (140+)",      "PASS — vào benchmark"),
        ("Gemma 2 2B IT",           "2B",    1.7,   "Có",             "Quá nhỏ cho writing task"),
        ("Llama 3.2 3B Instruct",   "3B",    2.0,   "Hạn chế",        "PASS — baseline (đang dùng)"),
        ("Phi-4-mini",              "3.8B",  2.5,   "KHÔNG",          "VRAM OK nhưng không hỗ trợ VN"),
        ("Arcee-VyLinh",            "3B",    2.0,   "Chuyên biệt VN", "VN chuyên biệt nhưng 3B quá nhỏ"),
    ]

    for i, (name, params, vram, vn, note) in enumerate(rows, start=5):
        hw_ok = vram <= 5.5
        vn_ok = vn.startswith("Có") or vn.startswith("Chuyên")
        pass_all = hw_ok and vn_ok and name not in ("Gemma 2 2B IT", "Arcee-VyLinh")

        values = [
            name, params, vram,
            "PASS" if hw_ok else "FAIL",
            vn,
            "PASS" if vn_ok else "FAIL",
            "✓ VÀO BENCHMARK" if pass_all else "Loại",
            note,
        ]
        write_row(ws, i, values)

        # Color coding
        ws.cell(row=i, column=4).fill = PASS_FILL if hw_ok else FAIL_FILL
        ws.cell(row=i, column=6).fill = PASS_FILL if vn_ok else FAIL_FILL
        result_cell = ws.cell(row=i, column=7)
        if pass_all:
            result_cell.fill = PASS_FILL
            result_cell.font = Font(bold=True, color="006100")
        else:
            result_cell.fill = FAIL_FILL

    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# SHEET 4 — COPY ORIGINAL GGUF TEST DATA
# ---------------------------------------------------------------------------
def build_sheet4_test_gguf(wb):
    """Copy original 'gguf test' sheet from Models.xlsx (preserve raw data)."""
    ws = wb.create_sheet("4_Test_GGUF_ban_dau")

    ws["A1"] = "GIAI ĐOẠN 4 — TEST BAN ĐẦU NHÓM GGUF (DATA GỐC TỪ NGHIÊN CỨU)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:M1")

    ws["A2"] = (
        "Test 5 model GGUF Q4_K_M với 5 prompt general (Kiến thức, Logic, Sáng tạo, "
        "Tóm tắt, Kỹ thuật). Dữ liệu được giữ nguyên từ file Models.xlsx gốc. "
        "Kết luận: Cần test lại với prompt SPECIFIC cho Report Agent (Sheet 5)."
    )
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 45

    # Copy data from original Excel
    try:
        old_wb = openpyxl.load_workbook(OLD_PATH, data_only=True)
        if "gguf test" in old_wb.sheetnames:
            old_ws = old_wb["gguf test"]
            for row_idx, row in enumerate(old_ws.iter_rows(values_only=True), start=4):
                if any(c is not None for c in row):
                    for col_idx, v in enumerate(row, start=1):
                        c = ws.cell(row=row_idx, column=col_idx, value=v)
                        c.border = BORDER
                        c.alignment = LEFT
                        if row_idx == 4:  # header row
                            c.fill = HEADER_FILL
                            c.font = HEADER_FONT
                            c.alignment = CENTER
            # Auto-size columns
            for col_idx in range(1, old_ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16
        else:
            ws["A4"] = "(Không tìm thấy sheet 'gguf test' trong Models.xlsx gốc)"
    except Exception as e:
        ws["A4"] = f"(Lỗi khi copy data: {e})"

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# SHEET 5 — REAL BENCHMARK
# ---------------------------------------------------------------------------
def build_sheet5_benchmark(wb):
    ws = wb.create_sheet("5_Benchmark_thuc_te")
    ws["A1"] = "GIAI ĐOẠN 5 — BENCHMARK THỰC TẾ VỚI PROMPT CỦA REPORT AGENT"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:K1")

    ws["A2"] = (
        "Test 4 candidate với 3 prompt thực tế từ LLMWriter: exec_summary (400 từ), "
        "remediation_detail (350 từ), recommendations (300 từ). Đo: tốc độ, VRAM, "
        "violations (word limit, first person, emoji, placeholder). "
        "Thiết bị: RTX 3060 Laptop 6GB VRAM. Nguồn: benchmark_results/benchmark_20260413_164412.json"
    )
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 55

    # === Sub-sheet 5A: Per-test metrics ===
    add_section_title(ws, 4, "A. Metrics chi tiết theo từng test", span=11)

    headers = [
        "Model", "Test", "Duration (s)", "Speed (tok/s)",
        "Words", "Word Limit", "VRAM (MB)", "Violations",
    ]
    widths = [32, 22, 14, 14, 10, 12, 12, 14]
    set_header(ws, 5, headers, widths)

    row = 6
    for model, data in BENCHMARK_DATA.items():
        is_chosen = model == "gemma3:4b"
        for test_name in ("exec_summary", "remediation_detail", "recommendations"):
            t = data[test_name]
            values = [
                model, test_name, t["duration"], t["speed"],
                t["words"], t["limit"], t["vram"], t["violations"],
            ]
            write_row(ws, row, values)

            # Highlight word limit violation
            if t["words"] > t["limit"] * 1.2:
                ws.cell(row=row, column=5).fill = FAIL_FILL
            else:
                ws.cell(row=row, column=5).fill = PASS_FILL

            if is_chosen:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
        row += 1  # blank row between models

    # === Sub-sheet 5B: Summary ===
    add_section_title(ws, row, "B. Tổng hợp hiệu năng", span=11)
    row += 1

    headers2 = [
        "Model", "Avg Speed (tok/s)", "Total Time (s)", "Max VRAM (MB)",
        "Total Violations", "Tốc độ vs baseline",
    ]
    widths2 = [32, 18, 16, 16, 16, 20]
    set_header(ws, row, headers2, widths2)
    row += 1

    baseline_speed = BENCHMARK_DATA["llama3.2:latest"]["summary"]["avg_speed"]
    for model, data in BENCHMARK_DATA.items():
        s = data["summary"]
        rel = s["avg_speed"] / baseline_speed
        values = [
            model, s["avg_speed"], s["total_time"], s["max_vram"],
            s["total_violations"], f"{rel:.2f}x",
        ]
        write_row(ws, row, values)
        if model == "gemma3:4b":
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PASS_FILL
                ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    row += 1

    # === Sub-sheet 5C: Quality scoring ===
    add_section_title(ws, row, "C. Chấm điểm chất lượng (manual review, thang 1-5)", span=11)
    row += 1

    headers3 = [
        "Model", "Tiếng Việt", "Instruction", "Accuracy",
        "Tone", "Concise", "TỔNG (25)",
    ]
    widths3 = [32, 14, 14, 14, 14, 14, 14]
    set_header(ws, row, headers3, widths3)
    row += 1

    for model, data in BENCHMARK_DATA.items():
        q = data["quality"]
        total = sum(q.values())
        values = [
            model, q["vietnamese"], q["instruction"], q["accuracy"],
            q["tone"], q["concise"], total,
        ]
        write_row(ws, row, values)
        total_cell = ws.cell(row=row, column=7)
        total_cell.font = Font(bold=True)
        if model == "gemma3:4b":
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PASS_FILL
                ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# SHEET 6 — DECISION MATRIX
# ---------------------------------------------------------------------------
def build_sheet6_quyet_dinh(wb):
    ws = wb.create_sheet("6_Quyet_dinh")
    ws["A1"] = "GIAI ĐOẠN 6 — MA TRẬN QUYẾT ĐỊNH"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:H1")

    ws["A2"] = (
        "Tổng hợp 6 tiêu chí với trọng số theo mức độ quan trọng cho Report Agent. "
        "Điểm thô 1-5, điểm có trọng số = điểm × trọng số / 20 (quy về thang 5)."
    )
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 35

    # Criteria with weights
    criteria = [
        # (name, weight_pct, scores_dict)
        ("Chất lượng tiếng Việt",   25, {"gemma3:4b": 5, "qwen2.5:7b-instruct-q4_K_M": 3, "deepseek-r1:7b": 2, "llama3.2:latest": 2}),
        ("Tốc độ (tok/s)",          20, {"gemma3:4b": 5, "qwen2.5:7b-instruct-q4_K_M": 2, "deepseek-r1:7b": 1, "llama3.2:latest": 5}),
        ("VRAM fit (6GB budget)",   15, {"gemma3:4b": 5, "qwen2.5:7b-instruct-q4_K_M": 4, "deepseek-r1:7b": 4, "llama3.2:latest": 5}),
        ("Instruction following",   15, {"gemma3:4b": 3, "qwen2.5:7b-instruct-q4_K_M": 5, "deepseek-r1:7b": 3, "llama3.2:latest": 1}),
        ("Độ chính xác nội dung",   15, {"gemma3:4b": 5, "qwen2.5:7b-instruct-q4_K_M": 5, "deepseek-r1:7b": 4, "llama3.2:latest": 2}),
        ("Văn phong chuyên nghiệp", 10, {"gemma3:4b": 5, "qwen2.5:7b-instruct-q4_K_M": 4, "deepseek-r1:7b": 4, "llama3.2:latest": 1}),
    ]

    models = ["gemma3:4b", "qwen2.5:7b-instruct-q4_K_M", "deepseek-r1:7b", "llama3.2:latest"]

    headers = ["Tiêu chí", "Trọng số (%)"] + models
    widths = [32, 14] + [28] * len(models)
    set_header(ws, 4, headers, widths)

    row = 5
    weighted_totals = {m: 0.0 for m in models}
    for crit_name, weight, scores in criteria:
        values = [crit_name, weight] + [scores[m] for m in models]
        write_row(ws, row, values)
        for m in models:
            weighted_totals[m] += scores[m] * weight / 100
        row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="TỔNG ĐIỂM (thang 5)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=2, value="100%").font = Font(bold=True)
    ws.cell(row=row, column=2).fill = SUBHEADER_FILL

    winner = max(weighted_totals, key=weighted_totals.get)
    for i, m in enumerate(models, start=3):
        c = ws.cell(row=row, column=i, value=round(weighted_totals[m], 2))
        c.font = Font(bold=True, size=12)
        c.alignment = CENTER
        c.border = BORDER
        if m == winner:
            c.fill = CHOSEN_FILL
            c.font = CHOSEN_FONT
        else:
            c.fill = NOTE_FILL

    # Rank row
    row += 1
    ranks = sorted(models, key=lambda m: weighted_totals[m], reverse=True)
    ws.cell(row=row, column=1, value="XẾP HẠNG").font = Font(bold=True)
    for i, m in enumerate(models, start=3):
        rank = ranks.index(m) + 1
        c = ws.cell(row=row, column=i, value=f"#{rank}")
        c.font = Font(bold=True)
        c.alignment = CENTER
        c.border = BORDER

    # === Decision justification ===
    row += 3
    ws.cell(row=row, column=1, value="QUYẾT ĐỊNH CUỐI").font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    justification = [
        ("Model được chọn:",    f"{winner}  (điểm: {round(weighted_totals[winner], 2)}/5.00)"),
        ("Lý do chính:",        "Cân bằng tốt nhất giữa chất lượng tiếng Việt (5/5), tốc độ (61.8 tok/s — nhanh thứ 2), và VRAM thoải mái (4.8GB/6GB)."),
        ("Ưu điểm:",            "Viết tiếng Việt tự nhiên nhất, văn phong chuyên nghiệp, tuân thủ cấu trúc."),
        ("Nhược điểm (chấp nhận):", "Hay vượt word limit (2/3 test) — cần tune num_predict hoặc constraint chi tiết hơn."),
        ("Phương án B:",        "Qwen2.5-7B-Instruct — instruction following tốt nhất (0 violations) nhưng output tiếng Anh và chậm 6x."),
        ("Phương án loại:",     "DeepSeek-R1 (chậm 9x, output tiếng Anh, thinking block tốn token); LLaMA 3.2 (dump prompt, VN không dấu)."),
    ]

    for label, content in justification:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2, value=content).alignment = LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# COPY ORIGINAL MODELS SHEET (preserve as backup)
# ---------------------------------------------------------------------------
def copy_original_as_backup(wb):
    """Copy original 'Models' sheet to '0_Original_backup' for audit trail."""
    try:
        old_wb = openpyxl.load_workbook(OLD_PATH, data_only=True)
        if "Models" not in old_wb.sheetnames:
            return
        old_ws = old_wb["Models"]
        ws = wb.create_sheet("0_Original_backup", 0)
        ws["A1"] = "BACKUP — Sheet 'Models' gốc từ Models.xlsx (giữ nguyên để audit)"
        ws["A1"].font = Font(bold=True, size=12, color="1F4E78")
        ws.merge_cells("A1:Q1")

        for row_idx, row in enumerate(old_ws.iter_rows(values_only=True), start=3):
            for col_idx, v in enumerate(row, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=v)
                c.border = BORDER
                c.alignment = LEFT
                if row_idx == 3:  # header
                    c.fill = HEADER_FILL
                    c.font = HEADER_FONT
        for col_idx in range(1, old_ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    except Exception as e:
        print(f"Warning: could not copy original: {e}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    copy_original_as_backup(wb)
    build_sheet1_khao_sat_llm(wb)
    build_sheet2_khao_sat_slm(wb)
    build_sheet3_loc_phan_cung(wb)
    build_sheet4_test_gguf(wb)
    build_sheet5_benchmark(wb)
    build_sheet6_quyet_dinh(wb)

    wb.save(OUT_PATH)
    print(f"[OK] Created: {OUT_PATH}")
    print(f"     Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
